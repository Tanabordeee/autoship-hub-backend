from sqlalchemy.orm import Session
from fastapi import UploadFile
import os
import base64
import requests
import io
import json
from pdf2image import convert_from_path
from app.schemas.lc import LCCreate
from app.repositories.lc_repo import LCRepo
from app.repositories.transaction_repo import TransactionRepo
from app.schemas.transaction import TransactionUpdate
from app.schemas.transaction import TransactionCreate
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session
from datetime import datetime
import re
def ocr_image(image, model: str):
    """Helper function to OCR a single image"""
    buf = io.BytesIO()
    image.save(buf, format="PNG")
    img_base64 = base64.b64encode(buf.getvalue()).decode()

    payload = {
        "model": model,
        "prompt": "อ่านข้อความทั้งหมดในภาพนี้อย่างละเอียด\n- รักษาลำดับบรรทัด\n- ไม่สรุป\n- ไม่ตีความ\n- พิมพ์ตามต้นฉบับ 100%",
        "images": [img_base64],
        "stream": False
    }

    r = requests.post(
        "http://localhost:11434/api/generate",
        json=payload
    )
    
    try:
        data = r.json()
        if "response" in data:
            return data["response"]
        else:
            print(f"Error: 'response' key missing. API returned: {data}")
            return f"[Error: {data.get('error', 'Unknown error')}]"
    except Exception as e:
        print(f"Exception during JSON parsing: {e}")
        print(f"Status Code: {r.status_code}")
        print(f"Raw Response: {r.text}")
        return "[Error: Failed to parse response]"
def create_lc(db:Session , payload: LCCreate , user_id:int , pi_id:list[int]):
    # Check if LC with same lc_no already exists
    existing_lc = LCRepo.get_latest_version_by_lc_no(db, payload.lc_no)
    # print(existing_lc.__dict__)
    if existing_lc:
        # Increment version
        new_version = (existing_lc.versions or 0) + 1
        
        # Merge None fields from new payload with existing LC data
        payload_dict = payload.model_dump()
        
        for field, value in payload_dict.items():
            if value is None and hasattr(existing_lc, field):
                # If new value is None, use the old value
                old_value = getattr(existing_lc, field)
                setattr(payload, field, old_value)
        
        # Set the new version number
        payload.versions = new_version
    else:
        # First version
        payload.versions = 1
    
    return LCRepo.create(db , payload , user_id , pi_id)
def clean_text_common(text: str) -> str:
    text = re.sub(
        r"THIS CREDIT IS VALID ONLY WHEN USED.*?(?=\n)|"
        r"NOTIFICATION OF LC ADVICE.*?(?=\n)|"
        r"PAGE \d+/.*?(?=\n)|"
        r"\[Page \d+\].*?(?=\n)|"
        r"^standard chartered\s*$|"
        r"^COMMERCIAL BANK OF CEYLON PLC\s*$|"
        r"^SH REL.*?(?=\n)|"
        r".*?ARR \.DATE=.*?(?=\n)|"
        r".*?ARR \.TIME=.*?(?=\n)|"
        r".*?REF \.NO\..*?(?=\n)|"
        r"^ARR .*?(?=\n)|"
        r"^REF .*?(?=\n)|"
        r"^DEAL=.*?(?=\n)|"
        r"^SENDER:.*?(?=\n)|"
        r".*?TEST AGREED SENDER.*?(?=\n)|"
        r"^Tel .*?(?=\n)|"
        r"^Fax .*?(?=\n)|"
        r"^Registration .*?(?=\n)|"
        r"^โทรศัพท์ .*?(?=\n)|"
        r"^โทรสาร .*?(?=\n)|"
        r"^ทะเบียนเลขที่ .*?(?=\n)|"
        r"^ธนาคารสแตนดาร์ดชาร์เตอร์ด.*?(?=\n)|"
        r"^140 ถนน.*?(?=\n)|"
        r"^140 Wireless.*?(?=\n)|"
        r"^ITSD-14.*?(?=\n)|"
        r"^\d+\s+TEST AGREED.*?(?=\n)|"
        r"^\d+\s*$|"
        r"^COLOMBO\s*$|"
        r"^Bangkok \d+.*?(?=\n)|"
        r"Standard Chartered Bank.*?(?=\n)|"
        r"TEST AGREED COMMERCIAL BANK OF CEYLON PLC COLOMBO.*?(?=\n)|"
        r"ICC PUBLICATION NO\.600 IS EXCLUDED.*?(?=\n)|"
        r"ARTICLE\s+\d+.*?UCP.*?(?=\n)|",
        "",
        text,
        flags=re.IGNORECASE | re.MULTILINE
    )

    text = re.sub(r"\n\s*\n+", "\n", text).strip()
    return text
def clean_45a_text(text: str) -> str:
    # ตัดทุกอย่างหลัง noise marker (STOP WORDS)
    stop_patterns = [
        r"THIS CREDIT IS VALID ONLY WHEN USED",
        r"NOTIFICATION OF LC ADVICE",
        r"PAGE\s+\d+/",
        r"\[Page\s*\d+\]",
        r"Standard Chartered Bank",
        r"ธนาคารสแตนดาร์ดชาร์เตอร์ด",
        r"COMMERCIAL BANK OF CEYLON",
        r"COMERCIAL BANK OF CEYLON",
        r"SH REL\. DATE",
        r"SENDER:",
    ]

    for p in stop_patterns:
        text = re.split(p, text, flags=re.IGNORECASE)[0]

    # cleanup whitespace
    text = re.sub(r"\n\s*\n+", "\n", text).strip()
    return text
def extract_document_require_46A(full_text: str):
    patterns = {
        1: r"46A\s*:\s*DOCUMENTS\s*REQUIRED\s*(.+?)(?=\s*2\))",
        2: r"2\)\s*(.+?)(?=\s*3\))",
        3: r"3\)\s*(.+?)(?=\s*4\))",
        4: r"4\)\s*(.+?)(?=\s*5\))",
        5: r"5\)\s*(.+?)(?=\s*6\))",
        6: r"6\)\s*(.+?)(?=\s*7\))",
        7: r"7\)\s*(.+?)(?=\s*:?\s*47A\s*:|$)",
    }

    doc_types = {
        1: "INVOICE",
        2: "BILL_OF_LADING",
        3: "INSURANCE",
        4: "CERTIFICATE_OF_REGISTRATION",
        5: "TRANSLATION",
        6: "INSPECTION_CERTIFICATE",
        7: "INSPECTION_CERTIFICATE",
    }

    items = []

    for item_no in range(1, 8):
        match = re.search(patterns[item_no], full_text, re.DOTALL | re.IGNORECASE)
        if not match:
            continue

        text = match.group(1).strip()

        # =================================================
        # CLEANUP LOGIC (ใช้ชุดเดิม 100% กับทุก item)
        # =================================================
        text = clean_text_common(text)

        text = re.sub(r"\n\s*\n+", "\n", text).strip()

        item = {
            "item_no": item_no,
            "doc_type": doc_types[item_no],
            "conditions": text
        }

        # =================================================
        # SPECIAL FORMAT : ITEM 6
        # =================================================
        if item_no == 6:

            annexures = []

            annexure_block_match = re.search(
                r"ORIGINAL\s+CERTIFICATE\s+OF\s+PRE\s+SHIPMENT\s+INSPECTION.*?"
                r"THIS\s+REPORT\s+SHOULD\s+HAVE\s+THE\s+FOLLOWING\s+ANNEXTURE\.(.+?)"
                r"(?=\n?\s*THE\s+STAMP\s+OF|\n?\s*\d+\)|$)",
                item["conditions"],
                flags=re.DOTALL | re.IGNORECASE
            )

            if annexure_block_match:
                annexure_block = annexure_block_match.group(1)

                annexure_matches = re.findall(
                    r"\(([A-C])\)\s*(.+?)(?=\n?\([A-C]\)|$)",
                    annexure_block,
                    flags=re.DOTALL | re.IGNORECASE
                )

                annexures = [
                    {
                        "code": code.upper(),
                        "text": re.sub(r"\n\s*\n+", "\n", body).strip()
                    }
                    for code, body in annexure_matches
                ]

            if annexures:
                item["annexures"] = annexures

            # ลบ annexure block ออกจาก conditions (ให้เหลือแต่ requirement หลัก)
            item["conditions"] = re.sub(
                r"THIS\s+REPORT\s+SHOULD\s+HAVE\s+THE\s+FOLLOWING\s+ANNEXTURE\..*",
                "",
                item["conditions"],
                flags=re.DOTALL | re.IGNORECASE
            ).strip()

        items.append(item)

    return {
        "items": items
    }
def extract_lc(db: Session, file: UploadFile, user_id: int , transaction_id:int):
    """
    Extract LC data from PDF file and return as JSON
    """
    # Save uploaded file
    upload_dir = "app/pdf"
    if not os.path.exists(upload_dir):
        os.makedirs(upload_dir)
    file_path = os.path.join(upload_dir, file.filename)
    with open(file_path, "wb") as buffer:
        buffer.write(file.file.read())

    PDF_PATH = file_path
    MODEL = "scb10x/typhoon-ocr1.5-3b:latest"
    POPPLER_PATH = r"E:\poppler\poppler-25.12.0\Library\bin"

    # Convert PDF to images (all pages)
    pages = convert_from_path(PDF_PATH, dpi=300, poppler_path=POPPLER_PATH)

    results = []
    for idx, page in enumerate(pages):
        print(f"OCR page {idx+1}/{len(pages)}")
        text = ocr_image(page, MODEL)
        results.append({"page": idx + 1, "text": text.strip()})

    # Combine all text
    full_text = "\n\n".join(f"[Page {r['page']}]\n{r['text']}" for r in results)

    # Extract all fields using regex
    extracted_data = {
        "sequence_of_total_27": re.search(r"SEQUENCE\s*OF\s*TOTAL\s*(.+?)(?=\s*:|$)", full_text, re.DOTALL|re.IGNORECASE),
        "form_of_documentary_credit_40a": re.search(r"FORM\s*OF\s*DOCUMENTARY\s*CREDIT\s*(.+?)(?=\s*:|$)", full_text, re.DOTALL|re.IGNORECASE),
        "docmentary_credit_number_20": re.search(r"DOCUMENTARY\s*CREDIT\s*NUMBER\s*(.+?)(?=\s*:|$)", full_text, re.DOTALL|re.IGNORECASE),
        "date_of_issue_31c": re.search(r"DATE\s*OF\s*ISSUE\s*(.+?)(?=\s*:|$)", full_text, re.DOTALL|re.IGNORECASE),
        "applicable_rules_40e": re.search(r"APPLICABLE\s*RULES\s*(.+?)(?=\s*:|$)", full_text, re.DOTALL|re.IGNORECASE),
        "date_and_place_of_expiry_31d": re.search(r"DATE\s*AND\s*PLACE\s*OF\s*EXPIRY\s*(.+?)(?=\s*:|$)", full_text, re.DOTALL|re.IGNORECASE),
        "applicant_50": re.search(r"APPLICANT\s*(.+?)(?=\s*:|$)", full_text, re.DOTALL|re.IGNORECASE),
        "beneficiary_59": re.search(r"59:\s*BENEFICIARY\s*(.+?)(?=\s*:|$)", full_text, re.DOTALL|re.IGNORECASE),
        "currency_code_32b": re.search(r"32B:\s*CURRENCY\s*CODE\s*,\s*AMOUNT\s*(.+?)(?=\s*:|$)", full_text, re.DOTALL|re.IGNORECASE),
        "available_with_41d": re.search(r"AVAILABLE\s*WITH\s*(.+?)(?=\s*:|$)", full_text, re.DOTALL|re.IGNORECASE),
        "partial_shipments_43p": re.search(r"PARTIAL\s*SHIPMENTS\s*(.+?)(?=\s*:|$)", full_text, re.DOTALL|re.IGNORECASE),
        "transhipment_43t": re.search(r"TRANSHIPMENT\s*(.+?)(?=\s*:|$)", full_text, re.DOTALL|re.IGNORECASE),
        "port_of_loading_of_departure_44e": re.search(r"PORT\s*OF\s*LOADING/AIRPORT\s*OF\s*DEPARTURE\s*(.+?)(?=\s*:|$)", full_text, re.DOTALL|re.IGNORECASE),
        "port_of_discharge_44f": re.search(r"PORT\s*OF\s*DISCHARGE/AIRPORT\s*OF\s*DESTINATION\s*(.+?)(?=\s*:|$)", full_text, re.DOTALL|re.IGNORECASE),
        "latest_date_of_shipment_44c": re.search(r"LATEST\s*DATE\s*OF\s*SHIPMENT\s*(.+?)(?=\s*:|$)", full_text, re.DOTALL|re.IGNORECASE),
        "additional_conditions_47a": re.search(r"47A\s*:\s*ADDITIONAL\s*CONDITIONS\s*(.+?)(?=\s*:|$)", full_text, re.DOTALL|re.IGNORECASE),
        "charges_71d": re.search(r"71D\s*:\s*CHARGES\s*(.+?)(?=\s*:|$)", full_text, re.DOTALL|re.IGNORECASE),
        "period_for_presentation_in_days_48": re.search(r"PERIOD\s*FOR\s*PRESENTATION\s*IN\s*DAYS\s*(.+?)(?=\s*:|$)", full_text, re.DOTALL|re.IGNORECASE),
        "confirmation_instructions_49": re.search(r"CONFIRMATION\s*INSTRUCTIONS\s*(.+?)(?=\s*:|$)", full_text, re.DOTALL|re.IGNORECASE),
        "instructions_to_the_paying_accepting_negotiating_bank_78": re.search(r"INSTRUCTIONS\s*TO\s*THE\s*PAYING/ACCEPTING/NEGOTIATING\s*BANK\s*(.*?)(?=THIS CREDIT IS VALID ONLY WHEN USED)", full_text, re.DOTALL|re.IGNORECASE),
        "lc_no": re.search(r"LC\s*ADVICE\s*NO.\s*(.+?)(?=\s*DATE)", full_text, re.DOTALL|re.IGNORECASE),
    }
    
    # Extract description of goods
    description_match = re.search(r"45A\s*:\s*DESCRIPTION\s*OF\s*GOODS\s*AND/OR\s*SERVICES\s*(.+?)(?=\s*\n?\s*\d{2}[A-Z]?\s*:|$)", full_text, re.DOTALL|re.IGNORECASE)

    # Build description_of_good_45a_45b as JSON with items
    description_of_good_45a_45b = None
    if description_match:
        raw_description_text = description_match.group(1).strip()
        description_text = clean_45a_text(raw_description_text)
        # Extract individual items (UNIT)
        items = re.split(
            r"(?=\b\d{1,2}\s+UNIT\b)",
            description_text,
            flags=re.IGNORECASE
        )
        items = [i.strip() for i in items if i.strip()]
        description_of_good_45a_45b = {
            "full_text": description_text,
            "items": [{"item_no": idx+1 , "description": item.strip()} for idx , item in enumerate(items)] if items else []
        }
    document_require_46a = extract_document_require_46A(full_text)
    # Build response JSON
    response_data = {
        "beneficiary_59": extracted_data["beneficiary_59"].group(1).strip() if extracted_data["beneficiary_59"] else None,
        "applicant_50": extracted_data["applicant_50"].group(1).strip() if extracted_data["applicant_50"] else None,
        "description_of_good_45a_45b": description_of_good_45a_45b,
        "date_of_issue_31c": extracted_data["date_of_issue_31c"].group(1).strip() if extracted_data["date_of_issue_31c"] else None,
        "lc_no": extracted_data["lc_no"].group(1).strip() if extracted_data["lc_no"] else None,
        "document_require_46a": document_require_46a if document_require_46a else None,
        "docmentary_credit_number_20": extracted_data["docmentary_credit_number_20"].group(1).strip() if extracted_data["docmentary_credit_number_20"] else None,
        "sequence_of_total_27": extracted_data["sequence_of_total_27"].group(1).strip() if extracted_data["sequence_of_total_27"] else None,
        "form_of_documentary_credit_40a": extracted_data["form_of_documentary_credit_40a"].group(1).strip() if extracted_data["form_of_documentary_credit_40a"] else None,
        "applicable_rules_40e": extracted_data["applicable_rules_40e"].group(1).strip() if extracted_data["applicable_rules_40e"] else None,
        "date_and_place_of_expiry_31d": extracted_data["date_and_place_of_expiry_31d"].group(1).strip() if extracted_data["date_and_place_of_expiry_31d"] else None,
        "currency_code_32b": extracted_data["currency_code_32b"].group(1).strip() if extracted_data["currency_code_32b"] else None,
        "available_with_41d": extracted_data["available_with_41d"].group(1).strip() if extracted_data["available_with_41d"] else None,
        "partial_shipments_43p": extracted_data["partial_shipments_43p"].group(1).strip() if extracted_data["partial_shipments_43p"] else None,
        "transhipment_43t": extracted_data["transhipment_43t"].group(1).strip() if extracted_data["transhipment_43t"] else None,
        "port_of_discharge_44f": extracted_data["port_of_discharge_44f"].group(1).strip() if extracted_data["port_of_discharge_44f"] else None,
        "port_of_loading_of_departure_44e": extracted_data["port_of_loading_of_departure_44e"].group(1).strip() if extracted_data["port_of_loading_of_departure_44e"] else None,
        "latest_date_of_shipment_44c": extracted_data["latest_date_of_shipment_44c"].group(1).strip() if extracted_data["latest_date_of_shipment_44c"] else None,
        "charges_71d": extracted_data["charges_71d"].group(1).strip() if extracted_data["charges_71d"] else None,
        "additional_conditions_47a": (
            clean_text_common(extracted_data["additional_conditions_47a"].group(1))
            if extracted_data["additional_conditions_47a"] else None
        ),
        "period_for_presentation_in_days_48": extracted_data["period_for_presentation_in_days_48"].group(1).strip() if extracted_data["period_for_presentation_in_days_48"] else None,
        "confirmation_instructions_49": extracted_data["confirmation_instructions_49"].group(1).strip() if extracted_data["confirmation_instructions_49"] else None,
        "instructions_to_the_paying_accepting_negotiating_bank_78": extracted_data["instructions_to_the_paying_accepting_negotiating_bank_78"].group(1).strip() if extracted_data["instructions_to_the_paying_accepting_negotiating_bank_78"] else None,
        "pdf_path": file_path,
        "full_text":full_text,
        "chunks": chunks
    }
    TransactionRepo.update(db, transaction_id , TransactionUpdate(status="pending", current_process="lc"))
    return response_data

def generate_excel(db: Session, id: int) -> str:
    lc = LCRepo.get_by_id(db, id)
    if not lc:
        raise Exception("LC not found")

    wb = Workbook()

    # =========================
    # Sheet 1 : LC HEADER
    # =========================
    ws = wb.active
    ws.title = "LC_HEADER"

    headers = [
        ("LC NO", lc.lc_no),
        ("Documentary Credit No", lc.docmentary_credit_number_20),
        ("Date of Issue", lc.date_of_issue_31c),
        ("Applicant", lc.applicant_50),
        ("Beneficiary", lc.beneficiary_59),
        ("Currency / Amount", lc.currency_code_32b),
        ("Port of Loading", lc.port_of_loading_of_departure_44e),
        ("Port of Discharge", lc.port_of_discharge_44f),
        ("Latest Shipment Date", lc.latest_date_of_shipment_44c),
        ("Created At", lc.created_at.replace(tzinfo=None) if lc.created_at else None),
    ]

    for row, (key, value) in enumerate(headers, start=1):
        ws.cell(row=row, column=1, value=key).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True)

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 80

    # =========================
    # Sheet 2 : GOODS 45A/45B
    # =========================
    ws_goods = wb.create_sheet("GOODS_45A")

    ws_goods.append(["Item No", "Description"])
    ws_goods["A1"].font = ws_goods["B1"].font = Font(bold=True)

    goods = lc.description_of_good_45a_45b or {}
    for item in goods.get("items", []):
        ws_goods.append([
            item.get("item_no"),
            item.get("description"),
        ])

    ws_goods.column_dimensions["A"].width = 15
    ws_goods.column_dimensions["B"].width = 120

    # =========================
    # Sheet 3 : DOCUMENTS 46A
    # =========================
    ws_docs = wb.create_sheet("DOC_REQUIRED_46A")

    ws_docs.append(["Item No", "Document Type", "Conditions"])
    for col in range(1, 4):
        ws_docs.cell(row=1, column=col).font = Font(bold=True)

    docs = lc.document_require_46a or {}
    for item in docs.get("items", []):
        ws_docs.append([
            item.get("item_no"),
            item.get("doc_type"),
            item.get("conditions"),
        ])

    ws_docs.column_dimensions["A"].width = 15
    ws_docs.column_dimensions["B"].width = 35
    ws_docs.column_dimensions["C"].width = 120

    # =========================
    # Sheet 4 : ANNEXURES
    # =========================
    ws_annex = wb.create_sheet("ANNEXURES")
    ws_annex.append(["Parent Item No", "Code", "Text"])
    for col in range(1, 4):
        ws_annex.cell(row=1, column=col).font = Font(bold=True)

    for item in docs.get("items", []):
        annexures = item.get("annexures", [])
        for annex in annexures:
            ws_annex.append([
                item.get("item_no"),
                annex.get("code"),
                annex.get("text"),
            ])

    ws_annex.column_dimensions["A"].width = 20
    ws_annex.column_dimensions["B"].width = 10
    ws_annex.column_dimensions["C"].width = 120

    # =========================
    # Save file
    # =========================
    os.makedirs("exports", exist_ok=True)
    file_path = f"exports/lc_{lc.id}.xlsx"
    wb.save(file_path)

    return file_path
