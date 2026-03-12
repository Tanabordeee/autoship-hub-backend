from sqlalchemy.orm import Session
from app.repositories.si import SI_Repository
from app.schemas.si import SICreate
from app.repositories.lc_repo import LCRepo
from app.repositories.booking import BookingRepo
from app.repositories.proforma_invoice_repo import ProformaInvoiceRepo
from app.repositories.vehicle_register import VehicleRegisterRepo
from app.repositories.transaction_repo import TransactionRepo
from app.schemas.transaction import TransactionUpdate
from weasyprint import HTML, CSS
import jinja2
import os
import re
from datetime import datetime
import logging
import num2words
from app.services.audit_log_service import audit_log_service
from openpyxl import load_workbook
from copy import copy
from app.repositories.proforma_invoice_repo import ProformaInvoiceRepo

logger = logging.getLogger(__name__)


def create_si(db: Session, payload: SICreate, user_id: int):
    pi = ProformaInvoiceRepo.get_pi_items_by_chassis(db, payload.chassis_no)
    si = SI_Repository.create_si(db, payload)
    lc = LCRepo.get_by_id(db, payload.lc_id)
    booking = BookingRepo.get_by_id(db, pi.booking_id)
    vehicle_register = VehicleRegisterRepo.get_by_id(db, pi.vehicle_register_id)
    proforma_invoice = ProformaInvoiceRepo.get_by_id(db, payload.pi_id)
    if not si or not lc or not booking or not vehicle_register or not proforma_invoice:
        return None

    logger.info("description_of_good_45a_45b: %s", lc.description_of_good_45a_45b)
    item_first_text = lc.document_require_46a["items"][0]["conditions"]
    match = re.search(
        r"(?:AS\s*)?PER\s*PROFORMA\s*INVOICE\s*NO\.?\s*[A-Z0-9-]+\s*OF\s*\d{1,2}\.\d{1,2}\.\d{4}",
        item_first_text,
    )
    as_per_proforma_invoice = match.group(0) if match else ""
    date_str = booking.etd
    dt = datetime.strptime(date_str, "%d/%m/%Y")
    etd = dt.strftime("%B %d, %Y").upper()
    number_of_original_bs = num2words.num2words(
        payload.number_of_original_bs, lang="en"
    ).upper()
    # Setup Jinja2
    template_path = r"E:\\job\\autoship-hub-server\\app\\templates\\si.html"
    template_dir = os.path.dirname(template_path)
    template_file = os.path.basename(template_path)

    env = jinja2.Environment(loader=jinja2.FileSystemLoader(template_dir))
    template = env.get_template(template_file)
    bank_text = lc.document_require_46a["items"][1]["conditions"]
    match = re.search(r"TO ORDER OF.*?SRI LANKA", bank_text, re.DOTALL)
    bank_lines = []
    if match:
        full_text = match.group(0)
        # ลบ comma ซ้ำ + เว้นวรรคเกิน
        full_text = re.sub(r"\s+", " ", full_text)

        # แยกด้วย comma
        parts = [p.strip() for p in full_text.split(",")]

        if len(parts) >= 3:
            bank_lines = [
                parts[0],  # TO ORDER OF COMMERCIAL BANK...
                ", ".join(parts[1:-1]),  # IMPORTS DEPT..., MAWATHA
                parts[-1],  # COLOMBO 01 SRI LANKA
            ]
    # Render HTML
    html_out = template.render(
        bank_lines=bank_lines,
        si=si,
        lc=lc,
        port_of_discharge=payload.port_of_discharge,
        port_of_loading=payload.port_of_loading,
        gross_weight=payload.gross_weight,
        measurement=payload.measurement,
        no_of_packages=payload.no_of_packages,
        number_of_original_bs=number_of_original_bs,
        original_bs=payload.number_of_original_bs,
        booking=booking,
        seal_no=payload.seal_no,
        vehicle_register=vehicle_register,
        proforma_invoice=proforma_invoice,
        as_per_proforma_invoice=as_per_proforma_invoice,
        container_no=payload.container_no,
        etd=etd,
    )
    # Convert to PDF using WeasyPrint
    output_path = payload.output_path
    if os.path.isdir(output_path):
        filename = f"si_{datetime.now().strftime('%Y%m%d%H%M%S')}.pdf"
        output_path = os.path.join(output_path, filename)

    css_path = os.path.join(os.getcwd(), "static", "tailwind.css")

    stylesheets = []
    if os.path.exists(css_path):
        stylesheets.append(CSS(filename=css_path))

    HTML(string=html_out).write_pdf(output_path, stylesheets=stylesheets)
    TransactionRepo.update(
        db,
        payload.transaction_id,
        TransactionUpdate(status="pending", current_process="si"),
        user_id=user_id,
    )
    audit_log_service.log_action(db, "create", "si", user_id, payload.transaction_id)
    ProformaInvoiceRepo.update_si_pi_items(db, payload.chassis_no, si.id)
    return {"output_path": output_path, "si_id": si.id}


def _copy_row_style(ws, src_row: int, dst_row: int, max_col: int = 10):
    for col in range(1, max_col + 1):
        s = ws.cell(row=src_row, column=col)
        d = ws.cell(row=dst_row, column=col)
        d._style = copy(s._style)
        d.font = copy(s.font)
        d.border = copy(s.border)
        d.fill = copy(s.fill)
        d.number_format = s.number_format
        d.protection = copy(s.protection)
        d.alignment = copy(s.alignment)
        d.value = None


def _shift_merges_down(ws, insert_at: int, amount: int):
    to_shift = []
    for rng in list(ws.merged_cells.ranges):
        min_c, min_r, max_c, max_r = rng.bounds
        if min_r >= insert_at:
            to_shift.append((min_c, min_r, max_c, max_r))

    for min_c, min_r, max_c, max_r in to_shift:
        ws.unmerge_cells(
            start_row=min_r, start_column=min_c, end_row=max_r, end_column=max_c
        )

    ws.insert_rows(insert_at, amount=amount)

    for min_c, min_r, max_c, max_r in to_shift:
        ws.merge_cells(
            start_row=min_r + amount,
            start_column=min_c,
            end_row=max_r + amount,
            end_column=max_c,
        )


def _set_merged_value(ws, row: int, col: int, value):
    """
    Safely set value for a cell, checking if it's part of a merge.
    If it is, set the value of the top-left cell of the merge range.
    """
    target_cell = ws.cell(row=row, column=col)

    # Check if the cell is part of any merged range
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            # It's merged. Write to the master cell.
            ws.cell(row=rng.min_row, column=rng.min_col).value = value
            return

    # Not merged, write directly.
    target_cell.value = value


def fill_data(ws, data):
    lc = data.get("lc")
    booking = data.get("booking")
    port_of_loading = data.get("port_of_loading")
    port_of_discharge = data.get("port_of_discharge")
    number_of_original_bs = data.get("number_of_original_bs")
    original_bs = data.get("original_bs")
    no_of_packages = data.get("no_of_packages")
    gross_weight = data.get("gross_weight")
    measurement = data.get("measurement")
    as_per_proforma_invoice = data.get("as_per_proforma_invoice")
    vehicle_register = data.get("vehicle_register")
    seal_no = data.get("seal_no")
    etd = data.get("etd")
    container_no = data.get("container_no")
    # ====== 1. ใส่ bank_lines เริ่มที่ A8 ======
    start_row = 8
    col = 1  # column A = 1
    bank_lines = data.get("bank_lines", [])
    for i, line in enumerate(bank_lines):
        _set_merged_value(ws, start_row + i, col, line)

    # ====== applicant_50 ลง A15 ไล่ลง ======
    start_row = 15
    col = 1  # A = 1
    applicant_text = lc.applicant_50 if lc and lc.applicant_50 else ""
    lines = applicant_text.splitlines()
    for i, line in enumerate(lines):
        _set_merged_value(ws, start_row + i, col, line)

    _set_merged_value(ws, 22, 3, lc.port_of_loading_of_departure_44e)  # C22
    _set_merged_value(ws, 25, 1, booking.feeder)  # A25
    _set_merged_value(ws, 25, 3, port_of_loading)  # C25
    _set_merged_value(ws, 27, 3, port_of_discharge)  # C27
    _set_merged_value(ws, 27, 1, port_of_discharge)  # A27
    _set_merged_value(ws, 27, 6, "BANGKOK")  # F27

    g_27 = number_of_original_bs + " " + (original_bs or "")
    _set_merged_value(ws, 27, 7, g_27)  # G27

    c_30_d_30 = str(no_of_packages or "") + " " + "UNIT"
    _set_merged_value(ws, 30, 3, c_30_d_30)  # C30
    _set_merged_value(ws, 30, 4, c_30_d_30)  # D30
    _set_merged_value(ws, 30, 6, gross_weight)  # F30
    _set_merged_value(ws, 30, 7, measurement)  # G30

    # ====== description_of_good ลง D31 ไล่ลง ======
    desc_data = lc.description_of_good_45a_45b if lc else None
    raw_desc_lines = []

    if isinstance(desc_data, list):
        for item in desc_data:
            if isinstance(item, dict):
                desc = item.get("description", "")
            else:
                desc = str(item)
            raw_desc_lines.extend(desc.splitlines())
    elif isinstance(desc_data, dict) and "items" in desc_data:
        for item in desc_data["items"]:
            desc = item.get("description", "")
            raw_desc_lines.extend(desc.splitlines())
    elif isinstance(desc_data, str):
        raw_desc_lines.extend(desc_data.splitlines())

    # Word wrap each line if it's too long
    all_desc_lines = []
    max_chars = 50  # Adjust as needed for column width
    for raw_line in raw_desc_lines:
        text = raw_line.strip()
        if not text:
            all_desc_lines.append("")
            continue

        while text:
            if len(text) <= max_chars:
                all_desc_lines.append(text)
                text = ""
            else:
                part = text[:max_chars]
                last_space = part.rfind(" ")
                if last_space != -1:
                    all_desc_lines.append(text[:last_space].strip())
                    text = text[last_space:].strip()
                else:
                    all_desc_lines.append(text[:max_chars].strip())
                    text = text[max_chars:].strip()

    # Template setup: Row 31 to 36 is for Description (6 rows)
    base_desc_rows = 6
    needed_rows = len(all_desc_lines)
    extra_rows = max(0, needed_rows - base_desc_rows)

    if extra_rows > 0:
        # Insert rows at 37
        _shift_merges_down(ws, 37, extra_rows)
        for i in range(extra_rows):
            _copy_row_style(ws, 31, 37 + i)

    # Fill description lines
    for i, line in enumerate(all_desc_lines):
        curr_row = 31 + i
        # Use helper for consistency
        _set_merged_value(ws, curr_row, 4, line)

        # Re-apply or handle merges for description
        # We only merge if they aren't already merged as requested
        # Note: we use merge_cells sparingly to avoid wiping borders
        if i == 0:
            # First line: Merge D:G (Col 4 to 7)
            # Check if already merged to avoid redundancy
            is_merged = any(
                rng.min_row == curr_row and rng.min_col == 4 and rng.max_col == 7
                for rng in ws.merged_cells.ranges
            )
            if not is_merged:
                ws.merge_cells(
                    start_row=curr_row, start_column=4, end_row=curr_row, end_column=7
                )
        else:
            # Subsequent lines: Merge D:E (Col 4 to 5)
            is_merged = any(
                rng.min_row == curr_row and rng.min_col == 4 and rng.max_col == 5
                for rng in ws.merged_cells.ranges
            )
            if not is_merged:
                ws.merge_cells(
                    start_row=curr_row, start_column=4, end_row=curr_row, end_column=5
                )

    # Dynamic row adjustments for footer
    offset = extra_rows
    _set_merged_value(
        ws,
        37 + offset,
        4,
        "CERTIFYING THAT VEHICLE IS " + (as_per_proforma_invoice or ""),
    )
    _set_merged_value(ws, 41 + offset, 4, vehicle_register.chassis_no)
    _set_merged_value(ws, 41 + offset, 6, vehicle_register.engine_no)

    credit_info = (
        "THE LETTER OF CREDIT NUMBER. "
        + (lc.docmentary_credit_number_20 or "")
        + " DATE OF ISSUE : "
        + (lc.date_of_issue_31c or "")
    )
    _set_merged_value(ws, 42 + offset, 4, credit_info)
    _set_merged_value(ws, 43 + offset, 4, lc.issuing_bank_52a)

    booking_seal = (container_no or "") + " / " + (seal_no or "")
    _set_merged_value(ws, 53 + offset, 1, booking_seal)

    shipped_info = (
        "SHIPPED ON BOARD DATE : " + (etd or "") + " AT " + (port_of_loading or "")
    )
    _set_merged_value(ws, 53 + offset, 4, shipped_info)
    _set_merged_value(ws, 54 + offset, 4, booking.feeder)
    _set_merged_value(ws, 59 + offset, 5, "BANGKOK THAILAND : " + (etd or ""))


def generate_excel(db, payload):
    template_path = r"E:\\job\\autoship-hub-server\\app\\templates\\si_template.xlsx"
    wb = load_workbook(template_path)
    ws = wb.active
    pi = ProformaInvoiceRepo.get_pi_items_by_chassis(db, payload.chassis_no)
    lc = LCRepo.get_by_id(db, payload.lc_id)
    booking = BookingRepo.get_by_id(db, pi.booking_id)
    vehicle_register = VehicleRegisterRepo.get_by_id(db, pi.vehicle_register_id)
    proforma_invoice = ProformaInvoiceRepo.get_by_id(db, payload.pi_id)
    if not lc or not booking or not vehicle_register or not proforma_invoice:
        return None
    logger.info("description_of_good_45a_45b: %s", lc.description_of_good_45a_45b)
    item_first_text = lc.document_require_46a["items"][0]["conditions"]
    match = re.search(
        r"(?:AS\s*)?PER\s*PROFORMA\s*INVOICE\s*NO\.?\s*[A-Z0-9-]+\s*OF\s*\d{1,2}\.\d{1,2}\.\d{4}",
        item_first_text,
    )
    as_per_proforma_invoice = match.group(0) if match else ""
    date_str = booking.etd
    dt = datetime.strptime(date_str, "%d/%m/%Y")
    etd = dt.strftime("%B %d, %Y").upper()
    number_of_original_bs = num2words.num2words(
        payload.number_of_original_bs, lang="en"
    ).upper()
    bank_text = lc.document_require_46a["items"][1]["conditions"]
    match = re.search(r"TO ORDER OF.*?SRI LANKA", bank_text, re.DOTALL)
    bank_lines = []
    if match:
        full_text = match.group(0)
        # ลบ comma ซ้ำ + เว้นวรรคเกิน
        full_text = re.sub(r"\s+", " ", full_text)

        # แยกด้วย comma
        parts = [p.strip() for p in full_text.split(",")]

        if len(parts) >= 3:
            bank_lines = [
                parts[0],  # TO ORDER OF COMMERCIAL BANK...
                ", ".join(parts[1:-1]),  # IMPORTS DEPT..., MAWATHA
                parts[-1],  # COLOMBO 01 SRI LANKA
            ]
    port_of_discharge = payload.port_of_discharge
    port_of_loading = payload.port_of_loading
    gross_weight = payload.gross_weight
    measurement = payload.measurement
    no_of_packages = payload.no_of_packages
    original_bs = payload.number_of_original_bs
    seal_no = payload.seal_no
    container_no = payload.container_no
    data = {
        "bank_lines": bank_lines,
        "number_of_original_bs": number_of_original_bs,
        "etd": etd,
        "as_per_proforma_invoice": as_per_proforma_invoice,
        "lc": lc,
        "booking": booking,
        "vehicle_register": vehicle_register,
        "proforma_invoice": proforma_invoice,
        "port_of_discharge": port_of_discharge,
        "port_of_loading": port_of_loading,
        "gross_weight": gross_weight,
        "measurement": measurement,
        "no_of_packages": no_of_packages,
        "original_bs": original_bs,
        "seal_no": seal_no,
        "container_no": container_no,
    }
    fill_data(ws, data)
    wb.save(payload.output_path)
    return {"output_path": payload.output_path}


def confirm_si(
    db: Session,
    transaction_id: int,
    si_id: int,
    user_id: int,
    image_base64: str | None = None,
):
    # Update SI with image if provided
    if image_base64:
        SI_Repository.update_si(db, si_id, image_base64)

    TransactionRepo.update(
        db,
        transaction_id,
        TransactionUpdate(status="completed", current_process="si", si_id=si_id),
        user_id=user_id,
    )
    audit_log_service.log_action(db, "confirm", "si", user_id, transaction_id)
    return True
