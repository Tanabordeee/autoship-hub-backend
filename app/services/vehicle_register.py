from sqlalchemy.orm import Session
import pdfplumber
import re
from app.repositories.transaction_repo import TransactionRepo
from app.schemas.transaction import TransactionUpdate
from app.schemas.vehicle_register import VehicleRegisterCreate
from app.repositories.vehicle_register import VehicleRegisterRepo
from app.repositories.proforma_invoice_repo import ProformaInvoiceRepo
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


def safe_search(pattern, content):
    match = re.search(pattern, content, re.MULTILINE | re.IGNORECASE | re.DOTALL)
    if match:
        return match.group(1).strip() if match.groups() else match.group(0).strip()
    return None


def extract_vehicle_register(db: Session, file, transaction_id: int):
    text = ""
    file.file.seek(0)
    with pdfplumber.open(file.file) as pdf:
        for page in pdf.pages:
            text += page.extract_text() or ""
    result = {
        "date_of_registration": safe_search(
            r"Date\s*of\s*Registration\s*(\d{1,2}\s+[A-Za-z]+\s+\d{4})", text
        ),  #
        "registration_no": safe_search(
            r"Registration\s*No.\s*(.*?)\s*(?=Province)", text
        ),  #
        "province": safe_search(r"Province\s*(.*?)(?=\s*Type)", text),  #
        "type_car": safe_search(r"Type\s*(.*?)(?=\s*Characteristics)", text),  #
        "characteristics": safe_search(r"Characteristics\s*:\s*([^\r\n]+)", text),  #
        "vehicle_make": safe_search(r"Vehicle\s*Make\s*:\s*(.*?)(?=\s*Model)", text),  #
        "model": safe_search(r"Model\s*:\s*(.*?)(?=\s*Model)", text),  #
        "model_year": safe_search(r"Model\s*year\s*(\d{4})", text),  #
        "colour": safe_search(r"Colour\s*(.*?)(?=\s*Chassis)", text),  #
        "chassis_no": safe_search(r"Chassis\s*No\.\s*(.*?)(?=\s+At\b)", text),  #
        "car_engine": safe_search(r"Car\s*Engine\s*(.*)(?=\s*Engine\s*No)", text),  #
        "engine_no": safe_search(r"Engine\s*No\.\s*(.*?)(?=\s*At\b)", text),  #
        "fuel_type": safe_search(r"Fuel\s*Type\s*(.*)(?=\s*Gas)", text),  #
        "vehicle_weight": safe_search(
            r"Vehicle\s*Weight\s*(.*)(?=\s*Loading)", text
        ),  #
        "total_weight": safe_search(r"Total\s*Weight\s*(.*)(?=\s*Seat)", text),  #
        "seat": safe_search(r"Seat\s*:\s*(\d+)", text),
    }
    TransactionRepo.update(
        db,
        transaction_id,
        TransactionUpdate(status="pending", current_process="vehicle_register"),
    )
    return result


def create_vehicle_register(
    db: Session, payload: VehicleRegisterCreate, user_id: int, transaction_id: int
):
    vehicle_register = VehicleRegisterRepo.create(db, payload, user_id, transaction_id)
    ProformaInvoiceRepo.update_vehicle_register_pi_items(
        db, payload.chassis, vehicle_register.id
    )
    TransactionRepo.update(
        db,
        transaction_id,
        TransactionUpdate(status="completed", current_process="vehicle_register"),
    )
    return {"id": vehicle_register.id}


def create_vehicle_register_excel(db: Session, id: int):
    vehicle_register = VehicleRegisterRepo.get_by_id(db, id)
    if not vehicle_register:
        raise ValueError("Vehicle Register not found")

    wb = Workbook()
    ws = wb.active
    ws.title = "Vehicle Register"

    headers = [
        ("date_of_registration", "Date of Registration"),
        ("registration_no", "Registration No"),
        ("province", "Province"),
        ("type_car", "Type Car"),
        ("characteristics", "Characteristics"),
        ("vehicle_make", "Vehicle Make"),
        ("model", "Model"),
        ("model_year", "Model Year"),
        ("colour", "Colour"),
        ("chassis_no", "Chassis No"),
        ("car_engine", "Car Engine"),
        ("engine_no", "Engine No"),
        ("fuel_type", "Fuel Type"),
        ("vehicle_weight", "Vehicle Weight"),
        ("total_weight", "Total Weight"),
        ("seat", "Seat"),
    ]

    # Header row
    for col, (_, title) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Data row
    for col, (field, _) in enumerate(headers, start=1):
        value = getattr(vehicle_register, field)
        ws.cell(row=2, column=col, value=value)

    # Auto width
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20

    file_path = f"/tmp/vehicle_register_{vehicle_register.id}.xlsx"
    wb.save(file_path)

    return file_path
