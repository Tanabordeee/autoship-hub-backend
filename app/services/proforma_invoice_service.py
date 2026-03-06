from sqlalchemy.orm import Session
from app.repositories.proforma_invoice_repo import ProformaInvoiceRepo
from app.schemas.proforma_invoice import CreateProformaInvoice
import os
import jinja2
from app.services.audit_log_service import audit_log_service
from weasyprint import HTML
from copy import copy
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment
from num2words import num2words
from openpyxl import load_workbook
from app.models.proforma_invoice import ProformaInvoice
from openpyxl.styles import Font
from openpyxl.styles import Border, Side


def create_proforma_invoice(db: Session, payload: CreateProformaInvoice, user_id: int):
    return ProformaInvoiceRepo.create(db, payload, user_id)


def generate_pdf(pi_id: str, db: Session, output_path: str):
    pi = ProformaInvoiceRepo.get_by_pi_id(db, pi_id)
    if not pi:
        raise Exception("Proforma Invoice not found")

    # Setup Jinja2
    template_path = r"E:\\job\\autoship-hub-server\\app\\templates\\invoice_weasy.html"
    template_dir = os.path.dirname(template_path)
    template_file = os.path.basename(template_path)
    # logo_path should be a file:// URL for WeasyPrint on Windows
    logo_file_path = r"E:\\job\\autoship-hub-server\\app\\assets\\logopap.png"
    logo_url = "file:///" + logo_file_path.replace("\\", "/")
    env = jinja2.Environment(loader=jinja2.FileSystemLoader(template_dir))
    template = env.get_template(template_file)
    # Render HTML
    html_out = template.render(invoice=pi, logo_path=logo_url)

    # Convert to PDF using WeasyPrint
    HTML(string=html_out).write_pdf(output_path)
    return output_path


def approve_proforma_invoice(db: Session, pi_id: str, approver: str, user_id: int):
    res = ProformaInvoiceRepo.update_pi_status(db, pi_id, "approved", approver)

    # Audit Log
    pi = ProformaInvoiceRepo.get_by_pi_id(db, pi_id)
    if pi and pi.transaction:
        audit_log_service.log_action(
            db, "approve", "proforma_invoice", user_id, pi.transaction.id
        )

    return res


def get_all_proforma_invoice(db: Session):
    return ProformaInvoiceRepo.get_all(db)


def reject_proforma_invoice(db: Session, pi_id: str, user_id: int):
    res = ProformaInvoiceRepo.update_pi_status(db, pi_id, "rejected")

    # Audit Log
    pi = ProformaInvoiceRepo.get_by_pi_id(db, pi_id)
    if pi and pi.transaction:
        audit_log_service.log_action(
            db, "reject", "proforma_invoice", user_id, pi.transaction.id
        )

    return res


def get_all_data_pi_items_by_pi_id_service(db: Session, pi_id: int):
    return ProformaInvoiceRepo.get_all_data_pi_items_by_pi_id(db, pi_id)


def get_proforma_invoice_by_pi_id(db: Session, pi_id: str):
    return ProformaInvoiceRepo.get_by_pi_id(db, pi_id)


def get_proforma_invoice_by_id(db: Session, id: int):
    return ProformaInvoiceRepo.get_by_id(db, id)


def get_chassis_by_pi_id(db: Session, pi_ids: int):
    row = ProformaInvoiceRepo.get_chassis_by_pi_id(db, pi_ids)
    return [r[0] for r in row]


def _fix_vertical_box(ws, row):
    thin = Side(style="thin")
    for col in [6, 7, 8]:  # F, G, H
        cell = ws.cell(row=row, column=col)
        cell.border = Border(
            left=thin,
            right=thin,
            top=cell.border.top,
            bottom=cell.border.bottom,
        )


# =====================================================
# UTILITIES (เหมือนของคุณ 100%)
# =====================================================


def _shift_merges_down(ws, insert_at: int, amount: int):
    to_shift = []
    for rng in list(ws.merged_cells.ranges):
        min_c, min_r, max_c, max_r = rng.bounds
        if min_r >= insert_at:
            to_shift.append((min_c, min_r, max_c, max_r))

    for min_c, min_r, max_c, max_r in to_shift:
        ws.unmerge_cells(
            start_row=min_r, start_column=min_c, end_row=max_r, end_column=max_c
        )

    ws.insert_rows(insert_at, amount=amount)

    for min_c, min_r, max_c, max_r in to_shift:
        ws.merge_cells(
            start_row=min_r + amount,
            start_column=min_c,
            end_row=max_r + amount,
            end_column=max_c,
        )


def _set_merged_value(ws, addr: str, value):
    cell = ws[addr]
    if not isinstance(cell, MergedCell):
        cell.value = value
        return

    r, c = cell.row, cell.column
    for rng in ws.merged_cells.ranges:
        min_c, min_r, max_c, max_r = rng.bounds
        if min_r <= r <= max_r and min_c <= c <= max_c:
            ws.cell(row=min_r, column=min_c).value = value
            return
    raise ValueError(f"MergedCell {addr} not found in merge range")


def _copy_row_style(ws, src_row: int, dst_row: int, max_col: int = 12):
    for col in range(1, max_col + 1):
        s = ws.cell(row=src_row, column=col)
        d = ws.cell(row=dst_row, column=col)
        d._style = copy(s._style)
        d.font = copy(s.font)
        d.border = copy(s.border)
        d.fill = copy(s.fill)
        d.number_format = s.number_format
        d.protection = copy(s.protection)
        d.alignment = copy(s.alignment)
        d.value = None


def _get_merges_in_row_range(ws, r1: int, r2: int):
    merges = []
    for rng in list(ws.merged_cells.ranges):
        min_c, min_r, max_c, max_r = rng.bounds
        if min_r >= r1 and max_r <= r2:
            merges.append((min_c, min_r, max_c, max_r))
    return merges


def _apply_merges_with_offset(ws, merges, row_offset: int):
    for min_c, min_r, max_c, max_r in merges:
        ws.merge_cells(
            start_row=min_r + row_offset,
            start_column=min_c,
            end_row=max_r + row_offset,
            end_column=max_c,
        )


def _ensure_term_rows(ws, term_count: int, offset: int):
    base_capacity = 5
    extra = max(0, term_count - base_capacity)
    if extra <= 0:
        return 0

    insert_at = 42 + offset
    _shift_merges_down(ws, insert_at=insert_at, amount=extra)

    for i in range(extra):
        _copy_row_style(ws, 37, insert_at + i)

    return extra


def set_long_text_across_cells(ws, start_cell: str, text: str):
    """
    ใส่ข้อความยาวแล้วตัดคำลงหลาย cell ด้านล่าง
    start_cell: เช่น "F14"
    text: ข้อความยาว
    """
    # cell แรก
    col = ws[start_cell].column  # column number
    row = ws[start_cell].row  # row number

    # ตัดข้อความเป็นบรรทัดยาวไม่เกิน 50 chars ต่อ cell (ปรับได้ตาม width)
    max_len = 50
    lines = []
    while text:
        part = text[:max_len]
        # ตัดที่ space ก่อนหน้านี้ ถ้าเป็นคำไม่ขาด
        if len(text) > max_len:
            last_space = part.rfind(" ")
            if last_space != -1:
                part = part[:last_space]
        lines.append(part.strip())
        text = text[len(part) :].strip()

    # ใส่ลง cell ต่อเนื่อง
    for i, line in enumerate(lines):
        ws.cell(row=row + i, column=col).value = line
        ws.cell(row=row + i, column=col).alignment = Alignment(vertical="top")


# =====================================================
# MAIN FUNCTION (ใช้ invoice แบบเดียวกับ HTML)
# =====================================================


def fill_pi_sheet(ws, invoice):

    # =========================
    # CONSIGNEE
    # =========================
    consignee_text = invoice.consignee_name or ""
    set_long_text_across_cells(ws, "A8", consignee_text)

    # =========================
    # NOTIFY PARTY
    # =========================
    notify_text = invoice.notify_party_name or ""
    set_long_text_across_cells(ws, "A14", notify_text)

    # =========================
    # SHIPPER
    # =========================
    shipper_text = invoice.shipper or ""
    set_long_text_across_cells(ws, "A20", shipper_text)

    # =========================
    # INVOICE BLOCK (RIGHT SIDE)
    # =========================
    _set_merged_value(ws, "F7", f"PRO-FORMA INVOICE  : {invoice.pi_id}")
    _set_merged_value(ws, "F8", f"DATE   :  {invoice.date}")
    _set_merged_value(ws, "F9", f"PAYMENT TERM  :  {invoice.payment_term}")
    _set_merged_value(ws, "F10", "PAP PROSPERITY CO LTD")
    _set_merged_value(ws, "F11", f"ACCOUNT NUMBER  {invoice.account_number}")
    _set_merged_value(ws, "F12", "CURRENT ACCOUNT")
    _set_merged_value(ws, "F13", f"SWIFT CODE  :  {invoice.swift_code}")
    bank_text = invoice.bank or ""
    set_long_text_across_cells(ws, "F14", bank_text)
    # =========================
    # GOODS (6 rows per vehicle — layout เดิมเป๊ะ)
    # =========================

    sorted_items = sorted(invoice.items, key=lambda x: x.item_no)

    goods_start = 27
    goods_rows = 6
    gap = 1
    block = goods_rows + gap

    # แยก items เป็นกลุ่มละ 6 เหมือน HTML
    vehicles = [sorted_items[i : i + 6] for i in range(0, len(sorted_items), 6)]

    offset_row = 0
    total_amount = 0

    if len(vehicles) > 1:
        extra_blocks = len(vehicles) - 1
        total_insert = extra_blocks * block
        insert_at = 33

        src_merges = _get_merges_in_row_range(
            ws, goods_start, goods_start + goods_rows - 1
        )
        _shift_merges_down(ws, insert_at, total_insert)

        offset_row += total_insert

        # Fix first block as well
        for i in range(goods_rows + gap):
            _fix_vertical_box(ws, goods_start + i)

        for k in range(extra_blocks):
            start_r = goods_start + (k + 1) * block
            for i in range(goods_rows + gap):
                _copy_row_style(ws, goods_start + i, start_r + i)
            _apply_merges_with_offset(ws, src_merges, (k + 1) * block)
            # Apply vertical borders to all rows in the block including the gap
            for i in range(goods_rows + gap):
                _fix_vertical_box(ws, start_r + i)

    # ใส่ข้อมูลทุกคัน
    for i, group in enumerate(vehicles):
        start_r = goods_start + i * block
        idx = i + 1

        desc = group[0].description if len(group) > 0 else ""
        year = group[1].description if len(group) > 1 else ""
        chassis = group[2].description if len(group) > 2 else ""
        hs_code = group[3].description if len(group) > 3 else ""

        freight = float(group[4].unit_price or 0) if len(group) > 4 else 0
        insurance = float(group[5].unit_price or 0) if len(group) > 5 else 0

        total_vehicle = float(group[0].amount_in_usd or 0)

        total_amount += total_vehicle

        _set_merged_value(ws, f"A{start_r}", idx)
        _set_merged_value(ws, f"B{start_r}", desc)
        _set_merged_value(ws, f"B{start_r + 1}", f"YEAR OF MANUFACTURE : {year}")
        _set_merged_value(ws, f"B{start_r + 2}", f"CHASSIS : {chassis}")

        _set_merged_value(ws, f"B{start_r + 3}", "FREIGHT")
        _set_merged_value(ws, f"F{start_r + 3}", freight)
        _set_merged_value(ws, f"G{start_r + 3}", 1)
        _set_merged_value(ws, f"H{start_r + 3}", f"=F{start_r + 3}*G{start_r + 3}")

        _set_merged_value(ws, f"B{start_r + 4}", "INSURANCE")
        _set_merged_value(ws, f"F{start_r + 4}", insurance)
        _set_merged_value(ws, f"G{start_r + 4}", 1)
        _set_merged_value(ws, f"H{start_r + 4}", f"=F{start_r + 4}*G{start_r + 4}")

        ws.merge_cells(
            start_row=start_r + 5, start_column=2, end_row=start_r + 5, end_column=3
        )
        _set_merged_value(ws, f"B{start_r + 5}", f"H.S. CODE {hs_code}")

        _set_merged_value(
            ws, f"F{start_r}", f"={total_vehicle}-F{start_r + 3}-F{start_r + 4}"
        )
        _set_merged_value(ws, f"G{start_r}", 1)
        _set_merged_value(ws, f"H{start_r}", f"=F{start_r}*G{start_r}")

    # =========================
    # TERMS
    # =========================
    insert_offset = offset_row
    if invoice.term_condition:
        # ใส่หัวข้อ TERM & CONDITION ที่ B43
        ws["B43"].value = "TERM & CONDITION"
        ws["B43"].font = Font(bold=True)
        thin_border = Border(bottom=Side(style="thin"))
        ws["B43"].border = thin_border

        terms = (invoice.term_condition or "").split("\n")
        insert_offset = _ensure_term_rows(ws, len(terms), offset_row) + offset_row
        for i, t in enumerate(terms):
            _set_merged_value(ws, f"B{37 + i + offset_row}", f"{i + 1}. {t}")
    # =========================
    # TOTAL WORDS
    # =========================
    words = num2words(int(round(total_amount)), lang="en").upper().replace("-", " ")

    _set_merged_value(ws, f"A{43 + insert_offset}", f"(TOTAL USD {words} ONLY)")

    ws[f"A{43 + insert_offset}"].alignment = Alignment(
        wrap_text=True, vertical="center", horizontal="center"
    )

    _set_merged_value(ws, f"H{44 + insert_offset}", f"=SUM(H26:H{42 + insert_offset})")

    # =========================
    # PORTS
    # =========================
    _set_merged_value(
        ws, f"A{46 + insert_offset}", f"PORT OF LOADING : {invoice.port_of_loading}"
    )
    ws[f"A{46 + insert_offset}"].font = Font(name="Arial", size=12)
    _set_merged_value(
        ws, f"A{45 + insert_offset}", f"PORT OF DISCHARGE : {invoice.port_of_discharge}"
    )
    ws[f"A{45 + insert_offset}"].font = Font(name="Arial", size=12)

    # =========================
    # SIGNATURE
    # =========================
    # Merge E48 ถึง G49
    ws.merge_cells(
        start_row=48 + insert_offset,
        start_column=5,  # E=5
        end_row=49 + insert_offset,
        end_column=7,
    )  # G=7

    # ใส่ข้อความ: ขีดบน + ชื่อ + OVERSEA SALES
    text = "_" * 30 + "\n" + f"{invoice.pi_approver}\nOVERSEA SALES"
    _set_merged_value(ws, f"E{48 + insert_offset}", text)

    # ตั้ง alignment
    ws[f"E{48 + insert_offset}"].alignment = Alignment(
        horizontal="center", vertical="top", wrap_text=True
    )

    customer_text = "_" * 30 + "\n" + f"{invoice.customer.customer_name}"
    _set_merged_value(ws, f"H{48 + insert_offset}", customer_text)

    ws[f"H{48 + insert_offset}"].alignment = Alignment(
        horizontal="center", vertical="top", wrap_text=True
    )
    # ลบค่า H49 เก่า
    ws[f"H{49 + insert_offset}"].value = None

    # Merge H48 กับ H49
    ws.merge_cells(
        start_row=48 + insert_offset,
        start_column=8,  # H = 8
        end_row=49 + insert_offset,
        end_column=8,
    )


def generate_excel(pi_id: str, db: Session, file_path: str):
    invoice: ProformaInvoice = (
        db.query(ProformaInvoice).filter(ProformaInvoice.pi_id == pi_id).first()
    )

    if not invoice:
        raise ValueError("Proforma Invoice not found")

    template_path = "E:\\job\\autoship-hub-server\\app\\templates\\pi_template.xlsx"

    wb = load_workbook(template_path)
    ws = wb.active

    fill_pi_sheet(ws, invoice)

    wb.save(file_path)
