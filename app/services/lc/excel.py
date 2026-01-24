import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from sqlalchemy.orm import Session
from app.repositories.lc_repo import LCRepo

def generate_excel(db: Session, id: int) -> str:
    lc = LCRepo.get_by_id(db, id)
    if not lc:
        raise Exception("LC not found")

    wb = Workbook()

    # =========================
    # Sheet 1 : LC_HEADER
    # =========================
    ws = wb.active
    ws.title = "LC_HEADER"

    headers = [
        ("LC NO", lc.lc_no),
        ("Documentary Credit No", lc.docmentary_credit_number_20),
        ("Date of Issue", lc.date_of_issue_31c),
        ("Applicant", lc.applicant_50),
        ("Beneficiary", lc.beneficiary_59),
        ("Latest Shipment Date", lc.latest_date_of_shipment_44c),
        ("Sequence of Total", lc.sequence_of_total_27),
        ("Form of Documentary Credit", lc.form_of_documentary_credit_40a),
        ("Applicable Rules", lc.applicable_rules_40e),
        ("Date and Place of Expiry", lc.date_and_place_of_expiry_31d),
        ("Currency / Amount", lc.currency_code_32b),
        ("Available With", lc.available_with_41d),
        ("Partial Shipments", lc.partial_shipments_43p),
        ("Transhipment", lc.transhipment_43t),
        ("Port of Loading", lc.port_of_loading_of_departure_44e),
        ("Port of Discharge", lc.port_of_discharge_44f),
        ("Charges", lc.charges_71d),
        ("Additional Conditions", lc.additional_conditions_47a),
        ("Period for Presentation in Days", lc.period_for_presentation_in_days_48),
        ("Confirmation Instructions", lc.confirmation_instructions_49),
        ("Instructions to the Paying Accepting Negotiating Bank", lc.instructions_to_the_paying_accepting_negotiating_bank_78),
    ]

    for row, (key, value) in enumerate(headers, start=1):
        ws.cell(row=row, column=1, value=key).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        ws.cell(row=row, column=2).alignment = Alignment(
            wrap_text=True, vertical="top"
        )

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 100

    # =========================
    # Sheet 2 : GOODS 45A/45B
    # =========================
    ws_goods = wb.create_sheet("GOODS_45A")

    ws_goods.append(["Item No", "Description"])
    ws_goods["A1"].font = ws_goods["B1"].font = Font(bold=True)

    goods = lc.description_of_good_45a_45b or {}
    for item in goods.get("items", []):
        ws_goods.append([
            item.get("item_no"),
            item.get("description"),
        ])

    for row in ws_goods.iter_rows(min_row=2, max_col=2):
        row[1].alignment = Alignment(wrap_text=True, vertical="top")

    ws_goods.column_dimensions["A"].width = 15
    ws_goods.column_dimensions["B"].width = 130

    # =========================
    # Sheet 3 : DOCUMENTS 46A
    # =========================
    ws_docs = wb.create_sheet("DOC_REQUIRED_46A")

    ws_docs.append(["Item No", "Document Type", "Conditions"])
    for col in range(1, 4):
        ws_docs.cell(row=1, column=col).font = Font(bold=True)

    docs = lc.document_require_46a or {}
    for item in docs.get("items", []):
        ws_docs.append([
            item.get("item_no"),
            item.get("doc_type"),
            item.get("conditions"),
        ])

    for row in ws_docs.iter_rows(min_row=2, max_col=3):
        row[2].alignment = Alignment(wrap_text=True, vertical="top")

    ws_docs.column_dimensions["A"].width = 15
    ws_docs.column_dimensions["B"].width = 35
    ws_docs.column_dimensions["C"].width = 130

    # =========================
    # Save file
    # =========================
    os.makedirs("exports", exist_ok=True)
    file_path = f"exports/lc_{lc.id}.xlsx"
    wb.save(file_path)

    return file_path
