from weasyprint import HTML, CSS
import jinja2
import os
from datetime import datetime
from app.schemas.commercial_invoice import CommercialInvoice
from app.repositories.proforma_invoice_repo import ProformaInvoiceRepo
from app.repositories.lc_repo import LCRepo
from app.repositories.bv import BVRepository
from app.repositories.si import SI_Repository
from app.repositories.booking import BookingRepo
import re
import ollama
from sqlalchemy.orm import Session
import json
from app.repositories.transaction_repo import TransactionRepo
from app.schemas.transaction import TransactionUpdate
from app.schemas.commercial_invoice import (
    CreateCommercialInvoicePayload,
)
from app.repositories.commercial_invoice import CommercialInvoiceRepo
from app.services.audit_log_service import audit_log_service
from openpyxl import load_workbook
from openpyxl.styles import Alignment


def call_gemma_extract(prompt: str, system_prompt: str):
    response = ollama.chat(
        model="qwen2.5:7b-instruct",
        messages=[
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": prompt},
        ],
    )
    return response["message"]["content"]


def parse_json_result(result_str: str):
    """
    Cleans and parses a JSON string from LLM output.
    Handles potential markdown code blocks.
    """
    if not result_str:
        return []
    try:
        # Check for markdown code blocks
        if "```json" in result_str:
            result_str = result_str.split("```json")[1].split("```")[0].strip()
        elif "```" in result_str:
            result_str = result_str.split("```")[1].split("```")[0].strip()

        # Clean up any potential leading/trailing non-json text
        result_str = result_str.strip()
        if not result_str.startswith("[") and "[" in result_str:
            result_str = result_str[result_str.find("[") :]
        if not result_str.endswith("]") and "]" in result_str:
            result_str = result_str[: result_str.rfind("]") + 1]

        return json.loads(result_str)
    except Exception as e:
        print(f"Error parsing JSON from LLM: {e}")
        # Fallback: if it's a multiline string that looks like a list
        lines = [
            line.strip().strip("-").strip("*").strip()
            for line in result_str.strip().split("\n")
            if line.strip()
        ]
        return lines


def generate_commercial_invoice(payload: CommercialInvoice, db: Session, user_id: int):
    pi = ProformaInvoiceRepo.get_by_id(db, payload.pi_id)
    lc = LCRepo.get_by_id(db, payload.lc_id)
    si = SI_Repository.get_by_id(db, payload.si_id)
    bv = BVRepository.get_by_id(db, payload.bv_id)
    booking = BookingRepo.get_by_id(db, payload.booking_id)
    # Setup Jinja2
    template_path = (
        r"E:\\job\\autoship-hub-server\\app\\templates\\commercial_invoice.html"
    )
    template_dir = os.path.dirname(template_path)
    template_file = os.path.basename(template_path)

    env = jinja2.Environment(loader=jinja2.FileSystemLoader(template_dir))
    template = env.get_template(template_file)

    # logo_path should be a file:// URL for WeasyPrint on Windows
    logo_file_path = r"E:\\job\\autoship-hub-server\\app\\assets\\logopap.png"
    logo_url = "file:///" + logo_file_path.replace("\\", "/")

    # Initialize variables for 46A extraction
    proforma_invoice_no = ""
    E_46A = []
    A_46A = ""
    beneficiary_to_certify = ""

    for i in (lc.document_require_46a or {}).get("items", []):
        if i.get("doc_type") == "INVOICE":
            conditions = i.get("conditions", "")
            match = re.search(r"PAP-\d+\s+OF\s+\d{2}\.\d{2}\.\d{4}", conditions)
            if match:
                proforma_invoice_no = match.group()
            E_46A_raw = call_gemma_extract(
                conditions,
                """
            You are a text extraction system.
From the input text below, extract ONLY the list of vehicle safety/accessory items
that are enumerated with (I), (II), (III), (IV).
Rules:
- Extract only the text belonging to items (I)–(IV)
- Do NOT include any other sentences
- Return the result as a JSON array of strings
- Keep the original wording exactly
            """,
            )
            print(f"DEBUG: conditions = {conditions}")
            E_46A = parse_json_result(E_46A_raw)
            print(f"DEBUG: E_46A = {E_46A}")
            A_46A = call_gemma_extract(
                conditions,
                """
You are a pure string slicing engine.

You are NOT allowed to interpret meaning.
You are NOT allowed to summarize.
You are NOT allowed to include any text outside the requested range.

TASK:

1. Locate the first exact occurrence of the characters: A)
2. Starting immediately AFTER those 2 characters,
3. Continue copying characters exactly as they appear.
4. Stop copying immediately BEFORE the first exact occurrence of: (B)
5. If "(B)" does not exist, copy until the end of the text.

STRICT RULES:

- Do not include "A)"
- Do not include "(B)"
- Do not skip characters
- Do not reformat
- Do not correct spelling
- Do not add or remove line breaks
- Do not explain anything
- Output raw text only

If you cannot find "A)", return empty.
            """,
            )
            beneficiary_to_certify = call_gemma_extract(
                conditions,
                """
You are a strict text extraction engine.

Extract ONLY the sentence that:
- Starts exactly with:
  "BENEFICIARY TO CERTIFY"
- Ends at the first period (.)

Rules:
- Use exact string matching
- Include the full sentence
- Do NOT include any text before it
- Do NOT include any text after the first period
- Preserve wording exactly
- Do NOT summarize
- Output plain text only
- No explanation
""",
            )
    desc_raw = lc.description_of_good_45a_45b or []
    if isinstance(desc_raw, dict):
        desc_items = desc_raw.get("items", [])
    elif isinstance(desc_raw, list):
        desc_items = desc_raw
    else:
        desc_items = []
    desc_text = desc_items[0].get("description", "") if desc_items else ""
    description = desc_text + (bv.engine_no if bv and bv.engine_no else "")
    freight_charge = 0
    insurance_charge = 0
    for p in pi.items:
        if p.description == "Freight":
            freight_charge = p.amount_in_usd
        if p.description == "Insurance":
            insurance_charge = p.amount_in_usd
    # Render HTML with dummy data
    html_out = template.render(
        invoice={
            "invoice_no": pi.pi_id,
            "date": pi.date,
            "proforma_invoice_no": proforma_invoice_no,
            "unit_price": pi.items[0].unit_price,
            "amount_in_usd": pi.items[0].amount_in_usd,
            "unit": pi.items[0].unit,
            "freight_charge": freight_charge,
            "insurance_charge": insurance_charge,
            "total_price": pi.total_price,
        },
        lc={
            "applicant": lc.applicant_50,
            "beneficiary": lc.beneficiary_59,
            "payment_terms": lc.form_of_documentary_credit_40a,
            "lc_number": lc.docmentary_credit_number_20,
            "date_of_issue": lc.date_of_issue_31c,
            "description": description,
            "E_46A": E_46A,
            "A_46A": A_46A,
            "beneficiary_to_certify": beneficiary_to_certify,
        },
        booking={
            "port_of_discharge": booking.port_of_disch,
            "port_of_loading": booking.port_of_loading,
        },
        si={
            "gross_weight": si.gross_weight,
            "port_of_loading": si.port_of_loading,
            "port_of_discharge": si.port_of_discharge,
        },
        bv={
            "country_of_origin": bv.country_of_origin,
            "year_of_manufacture": bv.year_of_manufacture,
            "make": bv.make,
            "model": bv.model,
            "type_of_vehicle": bv.type_of_vehicle,
            "year_of_registration": bv.year_month_of_first_registration,
            "bv_ref_no": bv.bv_ref_no,
            "commonly_called": bv.commonly_called,
        },
        bank=payload.bank,
        direction=payload.director,
        logo_path=logo_url,
    )

    # Ensure output directory exists
    output_dir = r"E:\\job\\autoship-hub-server\\app\\pdf"
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    # Generate filename
    filename = f"commercial_invoice_{datetime.now().strftime('%Y%m%d%H%M%S')}.pdf"
    output_path = os.path.join(output_dir, filename)

    # Add Tailwind CSS
    css_path = os.path.join(os.getcwd(), "static", "tailwind.css")
    stylesheets = []
    if os.path.exists(css_path):
        stylesheets.append(CSS(filename=css_path))

    # Convert to PDF using WeasyPrint
    # Providing base_url helps resolve relative assets if any
    HTML(string=html_out, base_url=template_dir).write_pdf(
        output_path, stylesheets=stylesheets
    )
    TransactionRepo.update(
        db,
        payload.transaction_id,
        TransactionUpdate(status="pending", current_process="commercial_invoice"),
        user_id=user_id,
    )
    audit_log_service.log_action(
        db, "generate", "commercial_invoice", user_id, payload.transaction_id
    )
    return {"output_path": output_path}


def _set_merged_value(ws, addr: str, value):
    cell = ws[addr]
    from openpyxl.cell.cell import MergedCell

    if not isinstance(cell, MergedCell):
        cell.value = value
        return

    r, c = cell.row, cell.column
    for rng in ws.merged_cells.ranges:
        min_c, min_r, max_c, max_r = rng.bounds
        if min_r <= r <= max_r and min_c <= c <= max_c:
            ws.cell(row=min_r, column=min_c).value = value
            return


def set_long_text_across_cells(
    ws, start_addr: str, end_row: int, text: str, max_len: int = 50
):
    """
    Fills text across multiple cells downward, wrapping at max_len.
    """
    if not text:
        return
    col_letter = "".join([c for c in start_addr if c.isalpha()])
    start_row = int("".join([c for c in start_addr if c.isdigit()]))

    lines = []
    text_to_process = str(text)
    while text_to_process:
        part = text_to_process[:max_len]
        if len(text_to_process) > max_len:
            last_space = part.rfind(" ")
            if last_space != -1:
                part = part[:last_space]
        lines.append(part.strip())
        text_to_process = text_to_process[len(part) :].strip()

    for i, line in enumerate(lines):
        if start_row + i > end_row:
            break
        _set_merged_value(ws, f"{col_letter}{start_row + i}", line)


def generate_commercial_invoice_excel(
    db: Session, payload: CommercialInvoice, user_id: int
):
    pi = ProformaInvoiceRepo.get_by_id(db, payload.pi_id)
    lc = LCRepo.get_by_id(db, payload.lc_id)
    si = SI_Repository.get_by_id(db, payload.si_id)
    bv = BVRepository.get_by_id(db, payload.bv_id)
    booking = BookingRepo.get_by_id(db, payload.booking_id)

    # Extraction logic (same as PDF version)
    proforma_invoice_no = ""
    E_46A = []
    A_46A = ""
    beneficiary_to_certify = ""

    for i in (lc.document_require_46a or {}).get("items", []):
        if i.get("doc_type") == "INVOICE":
            conditions = i.get("conditions", "")
            match = re.search(r"PAP-\d+\s+OF\s+\d{2}\.\d{2}\.\d{4}", conditions)
            if match:
                proforma_invoice_no = match.group()
            E_46A_raw = call_gemma_extract(
                conditions,
                """
            You are a text extraction system.
From the input text below, extract ONLY the list of vehicle safety/accessory items
that are enumerated with (I), (II), (III), (IV).
Rules:
- Extract only the text belonging to items (I)–(IV)
- Do NOT include any other sentences
- Return the result as a JSON array of strings
- Keep the original wording exactly
            """,
            )
            E_46A = parse_json_result(E_46A_raw)
            A_46A = call_gemma_extract(
                conditions,
                """
You are a pure string slicing engine.

You are NOT allowed to interpret meaning.
You are NOT allowed to summarize.
You are NOT allowed to include any text outside the requested range.

TASK:

1. Locate the first exact occurrence of the characters: A)
2. Starting immediately AFTER those 2 characters,
3. Continue copying characters exactly as they appear.
4. Stop copying immediately BEFORE the first exact occurrence of: (B)
5. If "(B)" does not exist, copy until the end of the text.

STRICT RULES:

- Do not include "A)"
- Do not include "(B)"
- Do not skip characters
- Do not reformat
- Do not correct spelling
- Do not add or remove line breaks
- Do not explain anything
- Output raw text only

If you cannot find "A)", return empty.
            """,
            )
            beneficiary_to_certify = call_gemma_extract(
                conditions,
                """
You are a strict text extraction engine.

Extract ONLY the sentence that:
- Starts exactly with:
  "BENEFICIARY TO CERTIFY"
- Ends at the first period (.)

Rules:
- Use exact string matching
- Include the full sentence
- Do NOT include any text before it
- Do NOT include any text after the first period
- Preserve wording exactly
- Do NOT summarize
- Output plain text only
- No explanation
""",
            )

    desc_raw = lc.description_of_good_45a_45b or []
    if isinstance(desc_raw, dict):
        desc_items = desc_raw.get("items", [])
    elif isinstance(desc_raw, list):
        desc_items = desc_raw
    else:
        desc_items = []
    desc_text = desc_items[0].get("description", "") if desc_items else ""
    description = desc_text

    # Load Template
    template_path = r"E:\\job\\autoship-hub-server\\app\\templates\\ci_template.xlsx"

    wb = load_workbook(template_path)
    ws = wb.active

    # 1. Applicant (A8-A10)
    set_long_text_across_cells(ws, "A8", 10, lc.applicant_50)

    # 2. Beneficiary (A14-A18)
    set_long_text_across_cells(ws, "A14", 18, lc.beneficiary_59)
    _set_merged_value(
        ws,
        "C1",
        "PAP Prosperity Co. Ltd  28 Soi Petchkasem 36, Pakkhlong Paseecharoen, Paseecharoen, Bangkok 10160, Thailand",
    )
    # 3. Fixed Right Side (F7-F14)
    _set_merged_value(ws, "F7", f"INVOICE : {pi.pi_id}")
    _set_merged_value(ws, "F8", f"DATE : {pi.date}")
    _set_merged_value(ws, "F9", f"PAYMENT TERMS : {lc.form_of_documentary_credit_40a}")
    _set_merged_value(
        ws, "F10", f"THE LETTER OF CREDIT NUMBER : {lc.docmentary_credit_number_20}"
    )
    _set_merged_value(ws, "F11", f"DATE OF ISSUE : {lc.date_of_issue_31c}")
    _set_merged_value(ws, "F12", "CERTIFYING THAT VEHICLE IS AS PER PROFORMA")
    _set_merged_value(ws, "F13", f"INVOICE NO {proforma_invoice_no}")
    _set_merged_value(ws, "F14", payload.bank)

    # 4. Item Row 23
    _set_merged_value(ws, "B23", description)
    _set_merged_value(ws, "F23", si.gross_weight)
    _set_merged_value(ws, "G23", pi.items[0].unit_price if pi.items else "")
    _set_merged_value(ws, "H23", pi.items[0].unit if pi.items else "")
    _set_merged_value(ws, "I23", pi.items[0].amount_in_usd if pi.items else "")

    # 5. Vehicle Details (B24-B30)
    origin_text = f"COUNTRY OF ORIGIN : {bv.country_of_origin}"
    manuf_text = f"YEAR OF MANUFACTURE : {bv.year_of_manufacture}"
    _set_merged_value(ws, "B24", origin_text)
    # Actually user said B24 for both. I'll put manuf in B25 if it's free, but user said MAKE is B25.
    # So I'll combine them in B24 with newline.
    _set_merged_value(ws, "B24", f"{origin_text}\n{manuf_text}")
    ws["B24"].alignment = Alignment(wrap_text=True)

    _set_merged_value(ws, "B25", f"MAKE : {bv.make}")
    _set_merged_value(ws, "B26", f"MODEL : {bv.model}")
    _set_merged_value(ws, "B27", f"TYPE OF VEHICLE : {bv.type_of_vehicle}")
    _set_merged_value(
        ws, "B28", f"YEAR OF FIRST REGISTRATION : {bv.year_month_of_first_registration}"
    )
    _set_merged_value(ws, "B29", f"EXPORT INSPECTION CERTIFICATE NO : {bv.bv_ref_no}")

    # 6. Dynamic 46A items starting B30
    curr_row = 30
    for item in E_46A:
        _set_merged_value(ws, f"B{curr_row}", item)
        ws[f"B{curr_row}"].alignment = Alignment(wrap_text=True)
        curr_row += 1

    if A_46A:
        _set_merged_value(ws, f"B{curr_row}", A_46A)
        ws[f"B{curr_row}"].alignment = Alignment(wrap_text=True)
        curr_row += 1

    if beneficiary_to_certify:
        _set_merged_value(ws, f"B{curr_row}", beneficiary_to_certify)
        ws[f"B{curr_row}"].alignment = Alignment(wrap_text=True)
        curr_row += 1

    # Calculate Charges
    freight_charge = 0
    insurance_charge = 0
    for p in pi.items:
        if p.description == "Freight":
            freight_charge = p.amount_in_usd
        if p.description == "Insurance":
            insurance_charge = p.amount_in_usd

    # 7. Totals & Footer
    _set_merged_value(ws, "I41", pi.items[0].amount_in_usd if pi.items else 0)

    commonly_called = bv.commonly_called if bv else ""
    pol = si.port_of_loading if si else ""
    pod = si.port_of_discharge if si else ""

    freight_text = f"FREIGHT CHARGE FOR 1 UNIT OF {commonly_called} FROM {pol} TO {pod}"
    _set_merged_value(ws, "A42", freight_text)
    _set_merged_value(ws, "I42", freight_charge)

    insurance_text = (
        f"INSRANCE CHARGE FOR 1 UNIT OF {commonly_called} FROM {pol} TO {pod}"
    )
    _set_merged_value(ws, "A43", insurance_text)
    _set_merged_value(ws, "I43", insurance_charge)

    _set_merged_value(ws, "H45", pi.items[0].unit if pi.items else "")
    _set_merged_value(ws, "I45", pi.total_price)

    gross_weight = si.gross_weight if si else 0
    _set_merged_value(
        ws, "A46", f"GRAND TOTAL NET WEIGHT & GROSS WEIGHT : {gross_weight} KGS"
    )

    # 8. Marks and Numbers
    _set_merged_value(ws, "F52", pi.items[0].unit if pi.items else "")
    _set_merged_value(ws, "F53", gross_weight)
    _set_merged_value(ws, "F54", gross_weight)

    # Save
    output_dir = r"E:\\job\\autoship-hub-server\\app\\excel"
    os.makedirs(output_dir, exist_ok=True)
    filename = f"commercial_invoice_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    output_path = os.path.join(output_dir, filename)
    wb.save(output_path)

    audit_log_service.log_action(
        db, "generate_excel", "commercial_invoice", user_id, payload.transaction_id
    )
    return {"output_path": output_path}


def confirm_commercial_invoice(
    db: Session, payload: CreateCommercialInvoicePayload, user_id: int
):
    commcercial_invoice = CommercialInvoiceRepo.create(db, payload, user_id)
    TransactionRepo.update(
        db,
        payload.transaction_id,
        TransactionUpdate(
            status="completed",
            current_process="commercial_invoice",
            commercial_invoice_id=commcercial_invoice.id,
        ),
        user_id=user_id,
    )
    audit_log_service.log_action(
        db, "confirm", "commercial_invoice", user_id, payload.transaction_id
    )
    return commcercial_invoice
