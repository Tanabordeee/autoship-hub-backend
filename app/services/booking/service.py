import pdfplumber
import re
from app.repositories.booking import BookingRepo
from app.schemas.booking import CreateBooking
from sqlalchemy.orm import Session
from app.repositories.transaction_repo import TransactionRepo
from app.repositories.proforma_invoice_repo import ProformaInvoiceRepo
from app.schemas.transaction import TransactionUpdate
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from app.services.audit_log_service import audit_log_service


def extract_booking(db: Session, file, transaction_id: int, user_id: int):

    text = ""
    file.file.seek(0)
    with pdfplumber.open(file.file) as pdf:
        for page in pdf.pages:
            text += page.extract_text() or ""

    def safe_search(pattern, content, group_idx=1):
        match = re.search(
            pattern,
            content,
            re.MULTILINE | re.IGNORECASE | re.DOTALL,  # <-- เพิ่ม DOTALL
        )
        if match:
            value = match.group(group_idx).strip()
            return re.sub(r"\s+", " ", value)  # ลบ newline ให้เหลือ space เดียว
        return None

    is_mercurial = bool(
        re.search(r"Mercurial\s+Logistics\s+Co\.,\s*Ltd\.", text, re.IGNORECASE)
    )

    is_renown = bool(
        re.search(r"RENOWN\s+TRANSPORT\s+CO\.,\s*LTD\.", text, re.IGNORECASE)
    )

    # ===============================
    # MERCURIAL TEMPLATE
    # ===============================
    if is_mercurial:
        result = {
            "date": safe_search(r"DATE\s*:\s*(\d{2}/\d{2}/\d{4})", text),
            "booking_no": safe_search(r"BOOKING\s*NO\.?\s*(\S+)", text),
            "carrier_booking_no": None,
            "carrier": safe_search(r"SHIPPING\s*LINE\s*(.*?)(?=\n|TAX)", text),
            "shipper": safe_search(r"SHIPPER\s*LOAD\s*(.*?)(?=\n)", text),
            "consignee": safe_search(r"TO\s*:\s*(.*?)(?=\n)", text),
            "fob_at": None,
            "quantity": safe_search(r"QUANTITIES\s*(.*?)(?=\n)", text),
            "feeder": safe_search(r"FEEDER\s*(.*?)(?=\n)", text),
            "vessel": safe_search(r"VESSEL\s*(.*?)(?=\n)", text),
            "place_of_rec": None,
            "port_of_loading": safe_search(
                r"PORT\s*OF\s*LOADING\s*(.*?)(?=\s*FREIGHT)", text
            ),
            "etd": safe_search(r"ETD\s*(\d{2}/\d{2}/\d{4})", text),
            "ts_port": safe_search(
                r"PORT\s*TRANSHIP\s*(.*?)(?=\nETD|\nCY|\nCLOSING)", text
            ),
            "port_of_disch": safe_search(r"DESTINATION\s*(.*?)(?=\n)", text),
            "eta_dest": safe_search(r"ETA\s*(\d{2}/\d{2}/\d{4})", text),
            "port_of_del": None,
            "final_destn": None,
            "cy_date": safe_search(r"CY\s*DATE\s*(\d{2}/\d{2}/\d{4})", text),
            "cy_at": safe_search(
                r"CY\s*AT/CTC\s*(.*?)(?=RETURN\s*AT/CTC)",
                text,
            ),
            "first_date_return": safe_search(
                r"First\s*return\s*(\d{2}/\d{2}/\d{4})", text
            ),
            "return_date": safe_search(r"RETURN\s*DATE\s*(\d{2}/\d{2}/\d{4})", text),
            "return_yard": safe_search(r"RETURN\s*AT/CTC\s*(.*?)(?=\nORDER)", text),
            "paperless_code": safe_search(r"PAPER\s*LESS\s*CODE\s*(\d+)", text),
            "closing_date": safe_search(r"CLOSING\s*TIME\s*(\d{2}/\d{2}/\d{4})", text),
            "at_before": safe_search(r"AT\s*BEFORE\s*(.*?)(?=\n)", text),
            "cut_off_si": safe_search(r"SI\s*CUT\s*OFF\s*(.*?)(?=\n)", text),
            "cut_off_vgm": safe_search(r"VGM\s*CUT\s*OFF\s*(.*?)(?=\n)", text),
            "booking_name": "Mercurial Logistics Co., Ltd.",
            "text": text,
        }
    elif is_renown:
        result = {
            "date": safe_search(r"BOOKING\s*CONFIRMATION\s*Date\s*([^\n]+)", text),
            "booking_no": safe_search(r"BKG\s*NO\.?\s*([A-Z0-9]+)", text),
            "carrier_booking_no": safe_search(
                r"Carrier\s*Booking\s*No\.?\s*([A-Z0-9]+)", text
            ),
            "carrier": safe_search(r"Carrier\s*(.*?)\s*Carrier\s*Tax", text),
            "shipper": safe_search(r"Shipper\s*(.*?)(?=\s*Commodity|\n)", text),
            "consignee": safe_search(r"TO\s*(.*?)(?=\s*Customer|\n)", text),
            "fob_at": None,
            "quantity": safe_search(r"Booking\s*Type\s*(.*?)\s*Booking\s*Date", text),
            "feeder": safe_search(r"Feeder\s*(.*?)\s*ETD", text),
            "vessel": safe_search(r"Vessel\s*(.*?)\s*ETA", text),
            "place_of_rec": safe_search(
                r"Place\s*of\s*Rec\.\s*(.*?)\s*Port\s*of\s*Loading", text
            ),
            "port_of_loading": safe_search(
                r"Port\s*of\s*Loading\s*(.*?)\s*T/S\s*Port", text
            ),
            "etd": safe_search(r"ETD[:\s]*([0-9\-\/]+)", text),
            "ts_port": safe_search(r"T/S\s*Port\s*(.*?)\s*Port\s*of\s*Discharge", text),
            "port_of_disch": safe_search(
                r"Port\s*of\s*Discharge\s*(.*?)\s*Final\s*Destination", text
            ),
            "eta_dest": safe_search(r"ETA\s*([0-9\/\-]+)", text),
            "port_of_del": None,
            "final_destn": safe_search(r"Final\s*Destination\s*(.*?)\s*Service", text),
            "cy_date": safe_search(r"CY\s*Date\s*([0-9\/\-]+)", text),
            "cy_at": safe_search(r"CY\s*Yard\s*(.*?)\s*Contact", text),
            "first_date_return": safe_search(
                r"First\s*Date\s*Return\s*w/o\s*Charge\s*([0-9\/\-]+)", text
            ),
            "return_date": safe_search(r"Return\s*Date\s*([0-9\/\-]+)", text),
            "return_yard": safe_search(r"Return\s*Yard\s*(.*?)\s*Contact", text),
            "paperless_code": safe_search(r"Paperless\s*Code\s*(\d+)", text),
            "closing_date": safe_search(
                r"Closing\s*Date\s*([0-9\/\-]+\s*[0-9:]+)", text
            ),
            "at_before": None,
            "cut_off_si": safe_search(r"SI\s*Cut-off\s*([0-9\/\-]+\s*[0-9:]+)", text),
            "cut_off_vgm": safe_search(r"VGM\s*Cut-off\s*([0-9\/\-]+\s*[0-9:]+)", text),
            "booking_name": "RENOWN TRANSPORT CO., LTD.",
            "text": text,
        }

    # ===============================
    # DEFAULT TEMPLATE (RIVA / Others)
    # ===============================
    else:
        result = {
            "date": safe_search(r"Date\s*:\s*(.*?)(?=\n|$)", text),
            "booking_no": safe_search(r"Booking\n?\s*No\.\s*:\s*(.*?)(?=\n|$)", text),
            "carrier_booking_no": safe_search(
                r"Carrier\s*Booking\s*No\s*:\s*(.*?)(?=\n|$)", text
            ),
            "carrier": safe_search(r"Carrier\s*:\s*(.*?)(?=\n|$)", text),
            "shipper": safe_search(r"Shipper\s*:\s*(.*?)(?=\n|$)", text),
            "consignee": safe_search(r"Consignee\s*:\s*(.*?)(?=\n|$)", text),
            "fob_at": safe_search(r"FOB\s*at\s*:\s*(.*?)(?=\n|$)", text),
            "quantity": safe_search(r"Quantities\s*:\s*(.*?)(?=\n|$)", text),
            "feeder": safe_search(r"Feeder\s*:\s*(.*?)(?=\n|$)", text),
            "vessel": safe_search(r"Vessel\s*:\s*(.*?)(?=\n|$)", text),
            "place_of_rec": safe_search(
                r"Place\s*of\s*Rec\.\s*:\s*(.*?)(?=\n|$)", text
            ),
            "port_of_loading": safe_search(
                r"Port\s*of\s*Loading\s*:\s*(.*?)(?=\s*ETD|$)", text
            ),
            "etd": safe_search(r"ETD\s*:\s*(.*?)(?=\n|$)", text),
            "ts_port": safe_search(r"T/S\s*Port\s*:\s*(.*?)(?=\n|$)", text),
            "port_of_disch": safe_search(
                r"Port\s*of\s*Disch\.\s*:\s*(.*?)(?=\s*ETA Dest|$)", text
            ),
            "eta_dest": safe_search(r"ETA\s*Dest\.\s*:\s*(.*?)(?=\n|$)", text),
            "port_of_del": safe_search(r"Port\s*of\s*Del\.\s*:\s*(.*?)(?=\n|$)", text),
            "final_destn": safe_search(r"Final\s*Destn\.\s*:\s*(.*?)(?=\n|$)", text),
            "cy_date": safe_search(r"CY\s*Date\s*:\s*(.*?)(?=\n|$)", text),
            "cy_at": safe_search(r"CY\s*AT\s*:\s*(.*?)(?=\n|$)", text),
            "first_date_return": safe_search(
                r"1st\s*Date\s*Return\s*:\s*(.*?)(?=\n|$)", text
            ),
            "return_date": safe_search(r"Return\s*Date\s*:\s*(.*?)(?=\n|$)", text),
            "return_yard": safe_search(r"Return\s*Yard\s*:\s*(.*?)(?=\n|$)", text),
            "paperless_code": safe_search(
                r"Paperless\s*Code\s*:\s*(.*?)(?=\n|$)", text
            ),
            "closing_date": safe_search(
                r"Closing\s*Date\s*:\s*(.*?)(?=\s*At Before|$)", text
            ),
            "at_before": safe_search(r"At\s*Before\s*:\s*(.*?)(?=\n|$)", text),
            "cut_off_si": safe_search(r"Cut\s*Off\s*SI\s*:\s*(.*?)(?=\n|$)", text),
            "cut_off_vgm": safe_search(r"Cut\s*Off\s*VGM\s*:\s*(.*?)(?=\n|$)", text),
            "booking_name": safe_search(r"(RIVA\s*LOGISTICS\s*CO\.,LTD\.)", text),
            "text": text,
        }

    TransactionRepo.update(
        db,
        transaction_id,
        TransactionUpdate(status="pending", current_process="booking"),
        user_id=user_id,
    )

    audit_log_service.log_action(db, "extract", "booking", user_id, transaction_id)

    return result


def create_booking(
    db: Session, payload: CreateBooking, user_id: int, transaction_id: int
):
    booking = BookingRepo.create(db, payload, user_id)
    ProformaInvoiceRepo.update_booking_pi_items(db, payload.chassis, booking.id)
    TransactionRepo.update(
        db,
        transaction_id,
        TransactionUpdate(
            status="completed", current_process="booking", booking_id=booking.id
        ),
        user_id=user_id,
    )
    audit_log_service.log_action(db, "create", "booking", user_id, transaction_id)
    return {"id": booking.id}


def create_booking_excel(db: Session, id: int, transaction_id: int, user_id: int):
    booking = BookingRepo.get_by_id(db, id)
    if not booking:
        raise ValueError("Booking not found")

    wb = Workbook()
    ws = wb.active
    ws.title = "Booking"

    headers = [
        ("booking_no", "Booking No"),
        ("booking_name", "Booking Name"),
        ("date", "Date"),
        ("carrier", "Carrier"),
        ("carrier_booking_no", "Carrier Booking No"),
        ("shipper", "Shipper"),
        ("consignee", "Consignee"),
        ("port_of_loading", "Port of Loading"),
        ("port_of_disch", "Port of Discharge"),
        ("port_of_del", "Port of Delivery"),
        ("etd", "ETD"),
        ("eta_dest", "ETA Destination"),
        ("cut_off_vgm", "Cut-off VGM"),
        ("cut_off_si", "Cut-off SI"),
        ("closing_date", "Closing Date"),
        ("return_date", "Return Date"),
        ("return_yard", "Return Yard"),
        ("cy_date", "CY Date"),
        ("cy_at", "CY At"),
        ("feeder", "Feeder"),
        ("place_of_rec", "Place of Receipt"),
        ("paperless_code", "Paperless Code"),
        ("fob_at", "FOB At"),
        ("quantity", "Quantity"),
    ]

    # Header row
    for col, (_, title) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Data row
    for col, (field, _) in enumerate(headers, start=1):
        value = getattr(booking, field)
        ws.cell(row=2, column=col, value=value)

    # Auto width
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20

    file_path = f"/tmp/booking_{booking.id}.xlsx"
    wb.save(file_path)
    audit_log_service.log_action(db, "export", "booking", user_id, transaction_id)
    return file_path

def get_booking_by_id(db : Session , id : int):
    return BookingRepo.get_by_id(db , id)